import io
import base64
import tempfile
from fastapi import APIRouter, Body, Request
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

router = APIRouter(prefix="/api/excel", tags=["excel"])


@router.post("/parse")
async def parse_excel(request: Request):
    try:
        raw_bytes = await request.body()
        if not raw_bytes or len(raw_bytes) < 20:
            return {"success": False, "message": "上传内容为空"}

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=False)
        wb_data = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_data = wb_data[sheet_name]
            rows = []
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column), 1):
                row_data = []
                for ci, cell in enumerate(row, 1):
                    formula = ""
                    computed_val = cell.value
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                        try:
                            computed_val = ws_data.cell(row=ri, column=ci).value
                        except Exception:
                            pass
                    row_data.append({
                        "v": str(computed_val) if computed_val is not None else "",
                        "f": formula
                    })
                    # 读取单元格格式
                    try:
                        if cell.font:
                            cell_style = {}
                            if cell.font.bold: cell_style["bold"] = True
                            if cell.font.italic: cell_style["italic"] = True
                            if cell.font.underline and cell.font.underline != 'none': cell_style["underline"] = True
                            if cell.font.strike: cell_style["strikethrough"] = True
                            if cell.font.size: cell_style["fontSize"] = cell.font.size
                            if cell.font.name: cell_style["fontFamily"] = cell.font.name
                            try:
                                if cell.font.color and cell.font.color.rgb:
                                    rgb = str(cell.font.color.rgb)
                                    if rgb and rgb != '00000000':
                                        cell_style["textColor"] = "#" + (rgb[-6:] if len(rgb) >= 6 else rgb)
                            except Exception:
                                pass
                            if cell_style:
                                row_data[-1].update(cell_style)
                        if cell.alignment and cell.alignment.horizontal:
                            row_data[-1]["textAlign"] = cell.alignment.horizontal
                        if cell.fill and cell.fill.start_color:
                            try:
                                rgb = str(cell.fill.start_color.rgb)
                                if rgb and rgb != '00000000':
                                    row_data[-1]["color"] = "#" + (rgb[-6:] if len(rgb) >= 6 else rgb)
                            except Exception:
                                pass
                    except Exception:
                        pass
                rows.append(row_data)
            sheets.append({
                "name": sheet_name,
                "rows": rows,
                "max_row": len(rows),
                "max_col": max(len(r) for r in rows) if rows else 0
            })
        wb.close()
        wb_data.close()
        return {"success": True, "data": {"sheets": sheets}}
    except Exception as e:
        return {"success": False, "message": f"Excel解析失败: {str(e)}"}


@router.post("/save")
async def save_excel(payload: dict = Body(...)):
    try:
        import openpyxl
        sheets_data = payload.get("sheets", [])
        if not sheets_data:
            return {"success": False, "message": "没有数据"}

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for sd in sheets_data:
            ws = wb.create_sheet(title=sd.get("name", "Sheet1"))
            rows = sd.get("rows", [])
            for ri, row in enumerate(rows, 1):
                for ci, cell_data in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci)
                    formula = cell_data.get("f", "")
                    value = cell_data.get("v", "")
                    if formula:
                        cell.value = formula
                    else:
                        cell.value = value
                    color = cell_data.get("color", "")
                    if color and color != '#ffffff' and color != '#000000':
                        cell.fill = PatternFill(
                            start_color=color.lstrip('#'),
                            end_color=color.lstrip('#'),
                            fill_type='solid'
                        )
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 写入单元格格式
                    bold = cell_data.get("bold", False)
                    italic = cell_data.get("italic", False)
                    underline = cell_data.get("underline", False)
                    strikethrough = cell_data.get("strikethrough", False)
                    font_size = cell_data.get("fontSize")
                    font_family = cell_data.get("fontFamily")
                    text_color = cell_data.get("textColor", "")
                    if any([bold, italic, underline, strikethrough, font_size, font_family, text_color]):
                        font_kwargs = {}
                        if bold: font_kwargs["bold"] = True
                        if italic: font_kwargs["italic"] = True
                        if underline: font_kwargs["underline"] = "single"
                        if strikethrough: font_kwargs["strike"] = True
                        if font_size: font_kwargs["size"] = font_size
                        if font_family: font_kwargs["name"] = font_family
                        if text_color: font_kwargs["color"] = text_color.lstrip('#')
                        if font_kwargs:
                            cell.font = Font(**font_kwargs)
                    text_align = cell_data.get("textAlign")
                    if text_align:
                        cell.alignment = Alignment(horizontal=text_align, vertical='center')

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        wb.close()
        with open(tmp.name, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('ascii')
        import os
        os.unlink(tmp.name)
        return {"success": True, "data_uri": f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}"}
    except Exception as e:
        return {"success": False, "message": f"Excel保存失败: {str(e)}"}
